import openpyxl
from plan.models import Plan, Purpose, Responsibl, Direction, Rubric, Plantable


def var_to_str(v):
    if v is None:
        return ''
    else:
        return str(v)


def var_to_int(v):
    if v is None:
        return 0
    else:
        return int(v)


def var_to_float(v):
    if v is None:
        return 0
    else:
        return float(v)


def start_import():  # Тут точка входу
    print('УВАГА!!!'
          'У папці d:/MyDoc/Dropbox/BASE/1/ мають розміщатися експортовані '
          'з IBExpert таблиці з іменами'
          'plan.xlsx, rozd.xlsx, meta.xlsx, napr.xlsx')

    pl_table = Plantable()
    pl_table.name = "asdfg"
    pl_table.save()
    pt_id = pl_table.id
    print("Створили Plantable з id=" + str(pt_id))

    rozd_ident, rozd_owner = imp_4(pt_id)

    purpose_ident = imp_2(pt_id)
    direction_ident = imp_3(pt_id)
    plan_ident, plan_rubric, plan_direction, plan_purpose = imp_1(pt_id)

    # Випраляємо посилання на id в Rubric
    rubrics = Rubric.objects.filter(plantable_id=pt_id)
    for rubric in rubrics:
        id_owner_old = rozd_owner[rubric.id]
        ident = rozd_ident[rubric.id]
        for key, value in rozd_ident.items():
            if value == id_owner_old:
                rubric.id_owner = Rubric.objects.get(pk=key)
                break

        #
        # per = Rubric.objects.filter(plantable_id=pt_id, ident=id_owner_old)[0]
        # i_id = per.id
        # rubric.id_owner = Rubric.objects.get(pk=i_id)
        rubric.save()

    
        
    
    # Випраляємо посилання в Plan
    plans = Plan.objects.filter(plantable_id=pt_id)
    for plan in plans:
        direction_old = plan_direction[plan.id]
        for key, value in  direction_ident.items():
            if value == direction_old:
                plan.direction_id = Direction.objects.get(pk=key)
                break


        purpose_old = plan_purpose[plan.id]
        for key, value in  purpose_ident.items():
            if value == purpose_old:
                plan.purpose_id = Purpose.objects.get(pk=key)
                break

        rubric_old = plan_rubric[plan.id]
        for key, value in rozd_ident.items():
            if value == rubric_old:
                plan.r_id = Rubric.objects.get(pk=key)
                break


        plan.save()
        


def imp_1(pt_id):
    print("Починаю експортувати")

    wb = openpyxl.load_workbook('d:/MyDoc/Dropbox/BASE/1/plan.xlsx')
    sheet = wb['Sheet1']
    # Рахуємо кількість заповнених рядків у таблиці Excel. У n номер останнього рядка з даними
    n = 1
    k = 1
    s = str(sheet['A' + str(n)].value)
    while k is not None:
        n += 1
        s = str(sheet['A' + str(n)].value)
        try:
            k = int(s)
        except:
            break
        # print(str(n)+'. '+s)
    # IDENT	ID_R	ZMIST	STROKI	FORMA	VIDPOV	PRIMITKA	SORT_	ID_NAPR	ID_META	TMP	SHOW

    plan_ident = {}
    plan_rubric = {}
    plan_direction = {}
    plan_purpose = {}

    for r in range(2, n):
        print(var_to_str(sheet['A' + str(r)].value))
        p = Plan()
        ident = var_to_str(sheet['A' + str(r)].value)
        rubric = int(var_to_str(sheet['B' + str(r)].value))

        p.content = var_to_str(sheet['C' + str(r)].value)
        p.termin = var_to_str(sheet['D' + str(r)].value)
        p.generalization = var_to_str(sheet['E' + str(r)].value)
        p.responsible = var_to_str(sheet['F' + str(r)].value)
        p.note = var_to_str(sheet['G' + str(r)].value)

        p.sort = var_to_float(sheet['H' + str(r)].value)

        direction = var_to_int(sheet['I' + str(r)].value)
        purpose = var_to_int(sheet['J' + str(r)].value)
        # p.tmp = var_to_str(sheet['K' + str(r)].value)
        p.plantable_id = Plantable.objects.get(pk=pt_id)
        p.show = True

        p.save()
        plan_ident[p.id] = ident
        plan_rubric[p.id] = rubric
        plan_direction[p.id] = direction
        plan_purpose[p.id] = purpose


    wb.close()

    print("Експорт завершено")
    return plan_ident, plan_rubric, plan_direction, plan_purpose


def imp_2(pt_id):
    print("Починаю експортувати")

    wb = openpyxl.load_workbook('d:/MyDoc/Dropbox/BASE/1/meta.xlsx')
    sheet = wb['Sheet1']
    # Рахуємо кількість заповнених рядків у таблиці Excel. У n номер останнього рядка з даними
    n = 1
    k = 1
    s = str(sheet['A' + str(n)].value)
    while k is not None:
        n += 1
        s = str(sheet['A' + str(n)].value)
        try:
            k = int(s)
        except:
            break
        # print(str(n)+'. '+s)
    purpose_ident = {}
    for r in range(2, n):
        print(var_to_str(sheet['A' + str(r)].value))
        p = Purpose()
        ident = var_to_str(sheet['A' + str(r)].value)
        p.name = var_to_str(sheet['B' + str(r)].value)
        p.plantable_id = Plantable.objects.get(pk=pt_id)
        p.save()
        purpose_ident[p.id] = ident

    wb.close()

    print("Експорт завершено")
    return purpose_ident



def imp_3(pt_id):
    print("Починаю експортувати")

    wb = openpyxl.load_workbook('d:/MyDoc/Dropbox/BASE/1/napr.xlsx')
    sheet = wb['Sheet1']
    # Рахуємо кількість заповнених рядків у таблиці Excel. У n номер останнього рядка з даними
    n = 1
    k = 1
    s = str(sheet['A' + str(n)].value)
    while k is not None:
        n += 1
        s = str(sheet['A' + str(n)].value)
        try:
            k = int(s)
        except:
            break
        # print(str(n)+'. '+s)
    direction_ident = {}
    for r in range(2, n):
        print(var_to_str(sheet['A' + str(r)].value))
        p = Direction()
        ident = var_to_int(sheet['A' + str(r)].value)
        p.name = var_to_str(sheet['B' + str(r)].value)
        p.plantable_id = Plantable.objects.get(pk=pt_id)

        p.save()
        direction_ident[p.id] = ident
    wb.close()

    print("Експорт завершено")
    return direction_ident


def imp_4(pt_id):
    print("Починаю експортувати")
    # TODO
    wb = openpyxl.load_workbook('d:/MyDoc/Dropbox/BASE/1/rozd.xlsx')
    sheet = wb['Sheet1']
    # Рахуємо кількість заповнених рядків у таблиці Excel. У n номер останнього рядка з даними
    n = 1
    k = 1
    s = str(sheet['A' + str(n)].value)
    while k is not None:
        n += 1
        s = str(sheet['A' + str(n)].value)
        try:
            k = int(s)
        except:
            break
        # print(str(n)+'. '+s)
    rozd_ident = {}
    rozd_owner = {}
    for r in range(2, n):
        print(var_to_str(sheet['A' + str(r)].value))
        p = Rubric()
        ident_tmp = var_to_int(sheet['A' + str(r)].value)
        # p.ident = var_to_int(sheet['A' + str(r)].value)

        p.n_r = var_to_int(sheet['B' + str(r)].value)
        p.name = var_to_str(sheet['C' + str(r)].value)
        owner_tmp = var_to_int(sheet['D' + str(r)].value)

        # p.ownertmp = var_to_int(sheet['D' + str(r)].value)
        p.plantable_id = Plantable.objects.get(pk=pt_id)
        p.riven = var_to_int(sheet['E' + str(r)].value)
        p.save()

        id = p.id
        rozd_ident[id] = ident_tmp
        rozd_owner[id] = owner_tmp
    wb.close()
    return rozd_ident, rozd_owner



    print("Експорт завершено")


"""
delete from plan_direction where true ;
delete from plan_plan where true ;
delete from plan_purpose where true ;
delete from plan_responsibl where true ;
delete from plan_rubric where true ;
delete from plan_terms where true ;


"""
