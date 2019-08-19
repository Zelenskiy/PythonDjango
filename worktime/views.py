import datetime
import os
from datetime import timedelta

import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Side

from django.core.files.storage import FileSystemStorage

from django.db.models import Q
from django.http import FileResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView
from openpyxl.styles import Border

from PythonDjango.settings import BASE_DIR, MEDIA_DIR
from timetable.models import Teacher, Timetable, Day, Card, Lesson, Subject, Resp, Class, Classroom, Period
from worktime.forms import SettingsForm, VacationForm, MissForm, HourlyForm
from worktime.models import Settings, Academyear, Vacat, Workday, Worktimetable, Missing, Hourlyworker


def str_to_datestr(ss1):
    mapping = [' ', ',', '.', '/', '-', '+', '*']
    for v in mapping:
        ss1 = ss1.replace(v, '.')
    ss1 = ss1.strip('.')
    while ss1.find('..') > -1:
        ss1 = ss1.replace('..', '.')
    n = ss1.count('.')
    if n == 1:
        ss1 = ss1 + '.' + str(datetime.datetime.today().year)
    return ss1


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days + 1)):
        yield start_date + timedelta(n)


def generatewd(request):
    # ay = Settings.objects.filter(field='academic_year')[0].value.strip()
    # workdays = Workday.objects.filter(acyear_id__name__iexact=ay)
    ttfrset = Settings.objects.filter(field='worktimetable')[0].value
    wt = Worktimetable.objects.get(pk=int(ttfrset))
    vacations = Vacat.objects.filter(worktimetable_id=wt, deleted=False)
    start = Settings.objects.filter(field='startacyear')[0].value.strip()
    end = Settings.objects.filter(field='endacyear')[0].value.strip()
    if request.POST:
        # вилучаємо робочі дні за поточний навчальний рік
        Workday.objects.filter(worktimetable_id=wt).delete()
        # Формуємо нову таблицю робочих днів
        d1 = datetime.datetime.strptime(start, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(end, '%d.%m.%Y')
        i = 0
        nw = 1
        n0 = 0
        for d in daterange(d1, d2):
            # print (d.strftime("%Y-%m-%d"))
            day = Workday()
            day.wday = d
            # Перевіряємо в таблиці Vacat
            count = len(Vacat.objects.filter(date=d))
            if count == 0:
                d_week = d.weekday() + 1
                day.dayweek = d_week
            else:
                d_week = 6
                day.dayweek = d.weekday() + 1

            if d_week == 6:
                day.numworkweek = 0
                day.num = 0
                day.weekchzn = 0
            elif d_week == 7:
                day.numworkweek = 0
                day.num = 0
                day.weekchzn = 0
                n0 = 1
            else:
                if n0 == 1:
                    nw += 1
                    n0 = 0
                day.numworkweek = nw
                i += 1
                day.num = i
                day.weekchzn = (nw + 1) % 2 + 1
            # print("wt="+str(wt))
            day.worktimetable_id = wt

            # day.acyear_id = Academyear.objects.get(pk=Vacat.objects.filter(worktimetable_id=wt)[0].id)
            day.save()
    ay = Settings.objects.filter(field='academic_year')[0].value.strip()

    workdays = Workday.objects.filter(worktimetable_id=wt)

    context = {'workdays': workdays, 'vacations': vacations, 'year': ay, 'start': start, 'end': end}
    return render(request, 'worktime/generatewd.html', context)


def index(request):
    ay = Settings.objects.filter(field='academic_year')[0].value.strip()
    ttfrset = Settings.objects.filter(field='worktimetable')[0].value
    wt = Worktimetable.objects.get(pk=int(ttfrset))
    workdays = Workday.objects.filter(worktimetable_id=wt)
    start = Settings.objects.filter(field='startacyear')[0].value.strip()
    end = Settings.objects.filter(field='endacyear')[0].value.strip()
    # vacations = Vacat.objects.filter(acyear_id__name__iexact=ay, deleted=False)
    if request.POST:
        for workday in workdays:
            s = request.POST['i-' + str(workday.id)]
            id_i = int(s)
            flag = False
            w = Workday.objects.get(pk=id_i)
            if (w.num != int(request.POST['n-' + s])):
                w.num = int(request.POST['n-' + s])
                flag = True
            dat_form = request.POST['w-' + s]
            dat_model = w.wday.strftime("%d.%m.%Y")

            if dat_model != dat_form:
                w.wday = datetime.datetime.strptime(request.POST['w-' + s], '%d.%m.%Y')
                flag = True
            if (w.numworkweek != int(request.POST['u-' + s])):
                w.numworkweek = int(request.POST['u-' + s])
                flag = True
            if (w.dayweek != int(request.POST['d-' + s])):
                w.dayweek = int(request.POST['d-' + s])
                flag = True
            if (w.weekchzn != int(request.POST['e-' + s])):
                w.weekchzn = int(request.POST['e-' + s])
                flag = True
            if flag:
                w.save()
                print("Збережено")
    # print(len(workdays))

    context = {'workdays': workdays, 'year': ay, 'start': start, 'end': end}

    return render(request, 'worktime/index.html', context)


class MissCreateView(CreateView):
    template_name = 'worktime/replace.html'
    form_class = MissForm
    success_url = '../../worktime/replace/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fill_context(context)
        # st = Settings.objects.filter(field='histreplst')[0].value.strip()
        # fin = Settings.objects.filter(field='histreplfin')[0].value.strip()
        # sthist = Settings.objects.filter(field='histzamst')[0].value.strip()
        # finhist = Settings.objects.filter(field='histzamfin')[0].value.strip()
        # fileHist = Settings.objects.filter(field='fileHist')[0].value.strip()
        # context['st'] = st
        # context['fin'] = fin
        # context['sthist'] = sthist
        # context['finhist'] = finhist
        # context['fileHist'] = fileHist
        return context

    def form_valid(self, form):
        ac = Academyear.objects.get(pk=1)
        # wt = Worktimetable.objects.filter(acyear_id=ac)[0]

        ttfrset = Settings.objects.filter(field='timetable')[0].value
        tt = Timetable.objects.get(pk=int(ttfrset))
        # tt = Timetable.objects.get(pk=int(Settings.objects.filter(field='timetable')[0].value))

        form.instance.timetable_id = tt
        return super(MissCreateView, self).form_valid(form)

    def form_invalid(self, form):
        print("invalid")
        return super(MissCreateView, self).form_invalid(form)


class HourlyCreateView(CreateView):
    template_name = 'worktime/setforpz.html'
    form_class = HourlyForm
    success_url = '../../worktime/setforpz/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ttfrset = Settings.objects.filter(field='timetable')[0].value
        tt = Timetable.objects.get(pk=int(ttfrset))
        context['teachers'] = Teacher.objects.filter(timetable_id=tt)

        return context

    def form_valid(self, form):
        ac = Academyear.objects.get(pk=1)
        # wt = Worktimetable.objects.filter(acyear_id=ac)[0]

        ttfrset = Settings.objects.filter(field='timetable')[0].value
        tt = Timetable.objects.get(pk=int(ttfrset))

        form.instance.timetable_id = tt
        return super(HourlyCreateView, self).form_valid(form)

    def form_invalid(self, form):
        print("invalid")
        return super(HourlyCreateView, self).form_invalid(form)


def set_border(ws, r1, c1, r2, c2):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for r in range(r1, r2):
        for c in range(c1, c2):
            ws.cell(row=r, column=c).border = thin_border


def psmaker(filename, d1, d2):
    # filename = 'd:/MyDoc/Dropbox/Школа/_Табелі, пояснюючі записки, розклад/2018-2019/Журнал замін 2018-2019.xlsx'
    # wb = openpyxl.load_workbook(filename)
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'media', filename))
    sheet = wb['Аркуш1']
    # Рахуємо кількість заповнених рядків у таблиці Excel. У n номер останнього рядка з даними
    n = 1

    s = str(sheet['A' + str(n)].value)
    history = []
    while s is not None:
        rec = {}
        n += 1
        s = str(sheet['A' + str(n)].value)
        dat = (sheet['B' + str(n)].value)
        # dat = datetime.datetime.strptime(datS,'%Y-%m-%d %h:%m:%s')
        # d1 = datetime.datetime.strptime('01.09.2019', '%d.%m.%Y')
        # d2 = datetime.datetime.strptime('30.09.2019', '%d.%m.%Y')

        try:
            k = int(s)
        except:
            break
        if (dat > d1 or dat == d1) and (dat < d2 or dat == d2):
            rec['D'] = dat

            rec['VT'] = str(sheet['C' + str(n)].value)
            rec['PV'] = str(sheet['D' + str(n)].value)
            rec['P1'] = str(sheet['E' + str(n)].value)
            rec['KL'] = str(sheet['F' + str(n)].value)
            rec['ZT'] = str(sheet['G' + str(n)].value)
            rec['P2'] = str(sheet['I' + str(n)].value)
            tmp = str(sheet['J' + str(n)].value)
            if tmp == 'pk':
                rec['poch_kl'] = True
            else:
                rec['poch_kl'] = False
            tmp = str(sheet['L' + str(n)].value)
            if tmp == 'kl_ker':
                rec['kl_ker'] = True
            else:
                rec['kl_ker'] = False
            history += [rec]
    print(n)

    teachers = []
    for h in history:
        # Визначаємо кількість вчителів
        title = h['VT'] + '~' + h['PV']
        reason = h['PV']
        # rec = {}
        if not title in teachers:
            teach = h['VT']
            subjects = []
            zamteachers = []
            rec = {}
            for h2 in history:
                title2 = h2['VT'] + '~' + h2['PV']
                rec4 = {}
                if title == title2:
                    subject = h2['P1']
                    if not subject in subjects:
                        klases = []
                        for h3 in history:
                            title3 = h3['VT'] + '~' + h3['PV']
                            if title2 == title3 and h3['P1'] == h2['P1']:
                                klas = h3['KL']
                                if not klas in klases:
                                    klases += [klas]
                        rec1 = {}
                        rec1['subject'] = subject
                        rec1['classes'] = klases
                        if not rec1 in subjects:
                            subjects += [rec1]

            rec['title'] = title
            rec['reason'] = reason
            rec['subjects'] = subjects
            rec['teach'] = teach
            # rec['zamteach'] = [rec3]
            if not rec in teachers:
                teachers += [rec]

    for t in teachers:
        count = 0
        date = []
        dss = []
        kl_ker = False
        poch_kl = False
        for h in history:
            title = h['VT'] + '~' + h['PV']

            if title == t['title']:
                count += 1
                ds = date_to_dd_mm(h['D'])
                if not ds in dss:
                    dss += [ds]
                date += [ds]
                if h['kl_ker']:
                    kl_ker = True
                if h['poch_kl']:
                    poch_kl = True

        t['kl_ker'] = (kl_ker)
        t['poch_kl'] = (poch_kl)
        cc, zzz = teach_to_zteach(history, t)
        t['zteachers'] = {'countZteach': cc, 'zts': zzz}

        t['count'] = count
        t['min'] = min_list(date)
        t['max'] = max_list(date)
        t['days'] = len(dss)

    listExcel = []

    for i, t in enumerate(teachers):
        rec = {}
        rec['N'] = str(i + 1)
        rec['teach'] = t['teach']
        subjects = t['subjects']
        s = ''
        for subject in subjects:
            klss = ''
            for kl in subject['classes']:
                klss += kl
            klss = klss.replace('-', '')
            s += subject['subject'] + '(' + klss + '),'
        rec['subject'] = s[:-1]
        rec['dmin'] = t['min']
        rec['dmax'] = t['max']
        rec['days'] = t['days']
        rec['count'] = t['count']
        rec['reason'] = t['reason']
        rec['zteachers'] = t['zteachers']
        rec['kl_ker'] = t['kl_ker']
        rec['poch_kl'] = t['poch_kl']

        listExcel += [rec]

    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'media', 'templ', 'PZap.xlsx'))
    sheet = wb['Аркуш1']

    n = 7
    for t in listExcel:
        sheet['B' + str(n)].value = t['N']
        sheet['C' + str(n)].value = t['teach']
        sheet['D' + str(n)].value = t['subject']
        sheet['E' + str(n)].value = t['dmin']
        sheet['F' + str(n)].value = t['dmax']
        sheet['G' + str(n)].value = t['days']
        sheet['H' + str(n)].value = t['count']
        sheet['L' + str(n)].value = t['reason']
        zts = t['zteachers']['zts']
        for zt in zts:
            z = zt['zteach']
            sheet['I' + str(n)].value = z
            countLesson = zt['countLesson']
            sheet['K' + str(n)].value = countLesson
            s = ''
            for subj in zt['subjects']:

                c = ''
                for cl in subj['classes']:
                    c += cl
                c = c.replace('-', '')
                s = subj['subject'] + '(' + c + ')'
                ss = s + '/' + str(subj['countKlas']) + 'год/'
                if ss[:3].upper() == 'ІНД':
                    ss += '+20%'
                sheet['J' + str(n)].value = ss
                n += 1

        n += 1
    num = len(listExcel)
    # Додаємо класне керівництво
    for t in listExcel:
        if t['kl_ker']:
            num += 1
            sheet['B' + str(n)].value = str(num)
            sheet['C' + str(n)].value = t['teach']
            sheet['D' + str(n)].value = 'Класне керівництво'
            sheet['E' + str(n)].value = t['dmin']
            sheet['F' + str(n)].value = t['dmax']
            sheet['G' + str(n)].value = t['days']

            sheet['L' + str(n)].value = t['reason']
            n += 2
            num += 1
    # Додаємо заміну ВГПД та вихователя з супроводу

    # Додаємо вивід погодинників
    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))
    hw = Hourlyworker.objects.filter(timetable_id=tt)
    for h in hw:

        subj = h.description

        sheet['B' + str(n)].value = str(num)
        sheet['C' + str(n)].value = h.teacher_id.short
        sheet['D' + str(n)].value = subj
        sheet['E' + str(n)].value = date_to_dd_mm(d1)
        sheet['F' + str(n)].value = date_to_dd_mm(d2)
        cards = h.teacher_id.cards.all()
        ttfrset = Settings.objects.filter(field='worktimetable')[0].value
        wt = Worktimetable.objects.get(pk=int(ttfrset))
        countLesson = 0
        workdays = set()
        for d in daterange(d1, d2):
            wdays = Workday.objects.filter(worktimetable_id=wt, wday=d)
            if len(wdays) < 1:
                continue
            wday = wdays[0]
            wdweek = wday.weekchzn  # Ціле число "1-чис., 2-зн., 0-вихідний"
            if wdweek == 1:
                wdw = '10'
            elif wdweek == 2:
                wdw = '01'
            else:
                continue
            # День тижня
            dayofweek = wday.dayweek - 1
            day = Day.objects.filter(timetable_id=tt, day=dayofweek)[0]

            for card in cards:
                lesson = card.lesson_id
                if (card.day_id == day) and ((lesson.weeks == '1') or (lesson.weeks == wdw)):

                    workdays.add(d)
                    periodsprcard = int(lesson.periodspercard)
                    for psc in range(periodsprcard):
                        countLesson += 1
        countDays = len(workdays)
        sheet['G' + str(n)].value = countLesson
        sheet['H' + str(n)].value = countDays

        num += 2
        n += 2

    # Додаємо вивід за пришкільні ділянки

    # Малюємо рамочку
    set_border(sheet, 7, 2, n - 1, 13)

    # Вивддимо дані підпису
    sheet['D' + str(n)].value = 'ЗД'
    sheet['I' + str(n)].value = Settings.objects.filter(field='Заступник директора')[0].value.upper()
    d01 = date_to_dd_mm(d1) + str(d1.year)
    d02 = date_to_dd_mm(d2) + str(d2.year)
    sheet['B4'].value = '(по хворобі чи інших обставинах) за період з ' + d01 + ' по ' + d02 + ''
    sheet['B1'].value = Settings.objects.filter(field='Школа (скорочено)')[0].value
    filename = os.path.join(BASE_DIR, 'media', 'report', 'PZap.xlsx')
    wb.save(filename)
    # Відправляємо файл на клієнт

    return filename


def teach_to_zteach(history, t):
    count = 0
    zteachers = []
    for h in history:
        title = t['title']
        htt = title.split('~')
        vt = htt[0]
        pv = htt[1]
        if h['VT'] == vt and h['PV'] == pv:
            zt = h['ZT']
            count += 1
            countSubj, sses = zteach_to_subjects(history, title, zt)
            if not sses in zteachers:
                zteachers += [sses]

    count = len(zteachers)
    return count, zteachers


def subject_to_classes(history, title, zt, sbs):
    classes = []
    count = 0
    for h in history:
        htt = title.split('~')
        vt = htt[0]
        pv = htt[1]
        if h['VT'] == vt and h['PV'] == pv and h['ZT'] == zt and h['P1'] == sbs:
            count += 1
            if not h['KL'] in classes:
                classes += [h['KL']]
    return count, classes


def zteach_to_subjects(history, title, zt):
    subjects = {'zteach': zt}
    subj = []
    count = 0
    for h in history:
        htt = title.split('~')
        vt = htt[0]
        pv = htt[1]
        if h['VT'] == vt and h['PV'] == pv and h['ZT'] == zt:
            count += 1
            if not h['P1'] in subj:
                subj += [h['P1']]
    subjects['countLesson'] = count
    pr = []
    for s in subj:
        dic = {}
        if not s in pr:
            c, clss = subject_to_classes(history, title, zt, s)
            dic['subject'] = s
            dic['classes'] = clss
            dic['countKlas'] = c
            pr += [dic]
    subjects['subjects'] = pr
    return count, subjects


def min_list(ls):
    min = ls[0]
    for l in ls:
        if l < min:
            min = l
    return min


def max_list(ls):
    max = ls[0]
    for l in ls:
        if l > max:
            max = l
    return max


def date_to_dd_mm(d):
    day = str(d.day)
    if len(day) == 1:
        day = '0' + day
    month = str(d.month)
    if len(month) == 1:
        month = '0' + month
    return day + '.' + month + '.'


def fill_context(context):
    st = Settings.objects.filter(field='histreplst')[0].value.strip()
    fin = Settings.objects.filter(field='histreplfin')[0].value.strip()
    sthist = Settings.objects.filter(field='histzamst')[0].value.strip()
    finhist = Settings.objects.filter(field='histzamfin')[0].value.strip()
    fileHist = Settings.objects.filter(field='fileHist')[0].value.strip()
    context['st'] = st
    context['fin'] = fin
    context['sthist'] = sthist
    context['finhist'] = finhist
    context['fileHist'] = fileHist


def prepeduplan(request):
    # Готуємо файл для введення навчального плану
    filename = os.path.join(BASE_DIR, 'media', 'templ', 'ENav.xlsx')
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb['Вчителі']
    sheet2 = wb['Предмети']

    sheet4 = wb['Класи']
    # sheet5 = wb['Кабінети']

    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))
    teachers = Teacher.objects.filter(timetable_id=tt)
    r = 1
    sheet1['A1'].value = ' -H- '
    for teacher in teachers:
        r += 1
        sheet1['A' + str(r)].value = teacher.name
    classes = Class.objects.filter(timetable_id=tt)
    r = 1
    for clas in classes:
        r += 1
        sheet4['A' + str(r)].value = clas.name
    # classrooms = Classroom.objects.filter(timetable_id=tt)
    # r=1
    # for clasr in classrooms:
    #     r+=1
    #     sheet5['A' + str(r)].value = clasr.name
    r = 1
    subjects = Subject.objects.filter(timetable_id=tt)
    for subj in subjects:
        r += 1
        sheet2['A' + str(r)].value = subj.short
    wb.save(filename)
    return FileResponse(open(filename, 'rb'), as_attachment=True)


def exptarif(request):
    # Експортуємо години для тарифікації
    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))
    teachers = Teacher.objects.filter(timetable_id=tt)

    text = ''
    for teacher in teachers:
        cl_1_4 = ('1-А')
        count_1_4 = 0
        count_5_9 = 0
        count_10_11 = 0
        cards = teacher.cards.all()
        for card in cards:
            lesson = card.lesson_id
            if lesson.weeks == '1':
                inc = 1
            elif lesson.weeks == '10' or lesson.weeks == '01':
                inc = 0.5
            else:
                inc = 0
            classes = lesson.classes.all()
            if len(classes) > 0:
                if (classes[0].name[:1] in ['1', '2', '3', '4']) and not (classes[0].name[1:2] in ['0', '1', '2']):
                    count_1_4 += inc
                elif classes[0].name[:1] in ['5', '6', '7', '8', '9']:
                    count_5_9 += inc
                elif classes[0].name[:1] == '1':
                    count_10_11 += inc

        text += (teacher.short + "\t" +
                 str(count_1_4).replace('.', ',') + "\t" +
                 str(count_5_9).replace('.', ',') + "\t" +
                 str(count_10_11).replace('.', ',')) + "\t" + "\n" + "\n"
        text = text.replace(',0', '').replace('\t0\t', '\t\t')

    filename = os.path.join(BASE_DIR, 'media', 'templ', 'exptarif.txt')
    open(filename, 'w').write(text)
    return FileResponse(open(filename, 'rb'), as_attachment=True)


def expfortab(request):
    # Експортуємо години для табеля
    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))
    teachers = Teacher.objects.filter(timetable_id=tt)
    cards = Card.objects.filter(timetable_id=tt)
    days = Day.objects.filter(timetable_id=tt)
    periods = Period.objects.filter(timetable_id=tt)
    tabel = []
    for teacher in teachers:
        rec = {}
        rec['id'] = teacher.id
        i = 0
        rec['name'] = teacher.name
        for w in range(2):
            for day in days:
               i += 1
               rec[i] = 0
        tabel += [rec]

    for card in cards:
        day = card.day_id
        lesson = card.lesson_id
        teachers = lesson.teachers.all()
        weeks = lesson.weeks
        mn = int(lesson.periodspercard)
        lesson_in_week = len(days)
        w1 = int(day.day) + 1
        w2 = lesson_in_week + w1
        for teacher in teachers:
            for tb in tabel:
                if tb['id'] == teacher.id:
                    if weeks == '10':
                        tb[w1] += mn
                    elif weeks == '01':
                        tb[w2] += mn
                    else:
                        tb[w1] += mn
                        tb[w2] += mn
                    break
    tex = ''
    for tb in tabel:
        tex += tb['name']+'\t'
        for d in range(1, 6):
            s = str(tb[d])
            if s=='0':
                s=''
            tex +=s+'\t'
        tex  += '\t\t'
        for d in range(6, 11):
            s = str(tb[d])
            if s == '0':
                s = ''
            tex += s + '\t'

        tex += '\n\n'

    filename = os.path.join(BASE_DIR, 'media', 'templ', 'expfortab.txt')
    open(filename, 'w').write(tex)
    return FileResponse(open(filename, 'rb'), as_attachment=True)


def expnavforcheck(request):
    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))

    cards = Card.objects.filter(timetable_id=tt)
    subjects = Subject.objects.filter(timetable_id=tt)
    classes = Class.objects.filter(timetable_id=tt)
    table = []
    for subject in subjects:
        rec = {}
        rec['subject'] = subject.id
        rec['subjectName'] = subject.name
        for clas in classes:
            rec[clas.name] = 0
        table += [rec]
    for card in cards:
        lesson = card.lesson_id
        weeks = lesson.weeks
        mn = int(lesson.periodspercard)
        if weeks == '1':
            h = 1 * mn
        else:
            h = 0.5 * mn
        subject = lesson.subjects.all()[0]
        classes = lesson.classes.all()
        for clas in classes:
            for s in table:
                if s['subject'] == subject.id:
                    s[clas.name] += h
                    break

    tex = 'Предмет' + "\t"
    subjects = Subject.objects.filter(timetable_id=tt)
    classes = Class.objects.filter(timetable_id=tt)
    for clas in classes:
        tex += clas.name + "\t"
    tex += "\n"
    for s in table:
        tex += s['subjectName'] + "\t"
        print(s['subjectName'])
        for clas in classes:
            ho = str(s[clas.name]).replace('.', ',')
            ho = ho.replace(' ', '')
            if ho == '0':
                ho = ''
            tex += ho + "\t"
        tex += "\n"
    # tex = tex.replace('\t0\t', '\t\t')

    filename = os.path.join(BASE_DIR, 'media', 'templ', 'expnavplan.txt')
    open(filename, 'w').write(tex)
    return FileResponse(open(filename, 'rb'), as_attachment=True)


def repl_3(request):
    if len(request.FILES) == 0:
        context = {}
        fill_context(context)
        return render(request, 'worktime/replace.html', context)
    if request.method == 'POST' and request.FILES['upload']:
        myfile = request.FILES['upload']

        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        s1 = request.POST['datepicker5']
        s2 = request.POST['datepicker6']
        d1 = datetime.datetime.strptime(s1, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(s2, '%d.%m.%Y')

        filename = psmaker(filename, d1, d2)
        # print('filename='+filename)
        # print('uploaded_file_url='+uploaded_file_url)
        # context['downloadpz'] = '<a href="{%  media "report/PZap.xlsx" %}" > Завантажити </a>'
    else:
        context = {}
        return render(request, 'worktime/replace.html', context)

    return FileResponse(open(filename, 'rb'), as_attachment=True)


@csrf_exempt
def datesave(request):
    if request.POST and request.is_ajax():
        d1 = request.POST['datepicker3']
        d2 = request.POST['datepicker4']
        s = Settings.objects.filter(field='histreplst')[0]
        s.value = d1
        s.save()
        s = Settings.objects.filter(field='histreplfin')[0]
        s.value = d2
        s.save()
    return render(request, 'worktime/replace.html', {})


@csrf_exempt
def pzdel(request, id):
    if request.POST and request.is_ajax():
        Hourlyworker.objects.get(pk=id).delete()
    return render(request, 'worktime/setforpz.html', {})


@csrf_exempt
def repldel(request, id):
    if request.POST and request.is_ajax():
        Missing.objects.get(pk=id).delete()
    return render(request, 'worktime/replace.html', {})


@csrf_exempt
def replgen(request):
    context = {}
    if request.POST and request.is_ajax():
        d1 = datetime.datetime.strptime(request.POST['datepicker3'], '%d.%m.%Y')
        d2 = datetime.datetime.strptime(request.POST['datepicker4'], '%d.%m.%Y')
        ttfrset = Settings.objects.filter(field='worktimetable')[0].value
        wt = Worktimetable.objects.get(pk=int(ttfrset))
        ttfrset = Settings.objects.filter(field='timetable')[0].value
        tt = Timetable.objects.get(pk=int(ttfrset))
        text = ''
        for d in daterange(d1, d2):
            wdays = Workday.objects.filter(worktimetable_id=wt, wday=d)
            if len(wdays) < 1:
                continue
            wday = wdays[0]
            wdweek = wday.weekchzn  # Ціле число "1-чис., 2-зн., 0-вихідний"
            if wdweek == 1:
                wdw = '10'
            elif wdweek == 2:
                wdw = '01'
            else:
                continue
            # День тижня
            dayofweek = wday.dayweek - 1
            q = Q(timetable_id=tt) & Q(date_st__lte=wday.wday) & Q(date_fin__gte=wday.wday)

            missings = Missing.objects.filter(q)
            day = Day.objects.filter(timetable_id=tt, day=dayofweek)[0]
            if len(missings) > 0:
                for missing in missings:
                    # tchmiss = missing.teach_id
                    teach = Teacher.objects.get(pk=missing.teach_id)
                    if teach != None:
                        cardsinteach = teach.cards.all()
                        # if len(cardsinteach) > 0:
                        for card in cardsinteach:
                            lesson = card.lesson_id
                            if (card.day_id == day) and ((lesson.weeks == '1') or (lesson.weeks == wdw)):
                                sbs = [x.short for x in lesson.subjects.all()]
                                ss = ''.join(sbs)
                                sbs = [x.short for x in lesson.classes.all()]
                                cs = ''.join(sbs)
                                reason = missing.reason
                                if reason == None:
                                    reason = ''
                                if missing.poch_kl:
                                    poch = 'pk'
                                else:
                                    poch = ''
                                periodsprcard = int(lesson.periodspercard)
                                for psc in range(periodsprcard):
                                    print(teach.short, "\t", d, "\t", ss, "\t", cs, "\t", reason)
                                    text += '\n\t' + d.strftime(
                                        "%d.%m.%y") + "\t" + teach.short + "\t" + reason + "\t" + ss + \
                                            "\t" + cs + "\t" + "\t" + "\t" + ss + "\t" + poch + "\t"
        resp = Resp()
        resp.timetable_id = tt
        resp.text = text
        resp.save()

        # Відпраляємо відповідь назад
        context = {}
    return render(request, 'worktime/replace.html', context)


def resp(request):
    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))
    r = Resp.objects.filter(timetable_id=tt)
    if len(r) > 0:
        context = {'resp': r[0].text[1:]}
    else:
        context = {}
    return render(request, 'worktime/resp.html', context)


def hwtable(request):
    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))
    hws = Hourlyworker.objects.filter(timetable_id=tt)
    context = {'hws': hws}
    return render(request, 'worktime/setforpzpnl.html', context)


def repltable(request):
    ttfrset = Settings.objects.filter(field='timetable')[0].value
    tt = Timetable.objects.get(pk=int(ttfrset))
    missing = Missing.objects.filter(timetable_id=tt)
    context = {'missing': missing}
    return render(request, 'worktime/repltable.html', context)


@csrf_exempt
def repladd(request):
    if request.POST and request.is_ajax():  # and request.user.has_perm('plan.change_plan'):
        r = Missing()

        # TODO виправити хрінь з цією табл
        r.worktimetable_id = Worktimetable.objects.get(pg=1)
        r.date_st = datetime.datetime.strptime(request.POST['datepicker1'], '%d.%m.%Y')
        r.date_fin = datetime.datetime.strptime(request.POST['datepicker2'], '%d.%m.%Y')
        r.reason = request.POST['reason']
        st = Settings.objects.filter(field='timetable')[0].value
        tt = Timetable.objects.get(pk=st)
        teachers = Teacher.objects.filter(timetable_id=tt)
        for teacher in teachers:
            pass

    return render(request, 'worktime/replace.html')


@csrf_exempt
def setchzn(request, id, chzn):
    workday = Workday.objects.get(pk=id)
    workday.weekchzn = chzn
    workday.save()
    return render(request, 'worktime/index.html')


def vacation(request):
    flag = 0
    form = VacationForm()
    wt = Settings.objects.filter(field='worktimetable')[0].value
    vacations = Vacat.objects.filter(worktimetable_id=wt, deleted=False)
    if request.POST:
        flag = 1
        for vac in vacations:
            s = request.POST['n_' + str(vac.id)]
            # існуючий запис
            id_i = int(s)
            v = Vacat.objects.get(pk=id_i)
            v.date = datetime.datetime.strptime(request.POST['dat_' + s], '%d.%m.%Y')
            v.name = request.POST['nam_' + s]
            if request.POST.get('del_' + s, False):
                v.deleted = True
            v.save()
        # додаємо нові
        for n in range(1, 5):
            ss1 = request.POST.get('dat-' + str(n), '')
            ss2 = request.POST.get('nam-' + str(n), '')
            if ss1 != '' and ss2 != '':
                v = Vacat()
                v.name = ss2
                ss1 = str_to_datestr(ss1)
                try:
                    v.date = datetime.datetime.strptime(ss1, '%d.%m.%Y')
                except:
                    try:
                        v.date = datetime.datetime.strptime(ss1, '%d.%m.%y')
                    except:
                        print("Формат дати невірний")
                v.worktimetable_id = Worktimetable.objects.get(pk=wt)
                # v.acyear_id = Academyear.objects.get(pk=Vacat.objects.filter(acyear_id__name__iexact=ay)[0].id)
                v.save()
        # ay = Settings.objects.filter(field='academic_year')[0].value.strip()
        vacations = Vacat.objects.filter(worktimetable_id=wt, deleted=False)
        Vacat.objects.filter(deleted=True).delete()

    context = {'vacations': vacations, 'form': form, 'flag': flag}
    return render(request, 'worktime/vacations.html', context)


def setting(request):
    settings = Settings.objects.all()
    kr = []
    flag = 0
    for set in settings:
        kr.append(set.field)
    form = SettingsForm()
    if request.POST:
        flag = 1
        for set in settings:
            f = str(request.POST['field_' + set.field]).strip()
            v = str(request.POST[set.field]).strip()
            for x in kr:
                if (set.field == x) and (f == x):
                    set.value = v
            set.save()
            print("Збережено ")
    context = {'form': form, 'settings': settings, 'flag': flag}
    return render(request, 'worktime/settings.html', context)
